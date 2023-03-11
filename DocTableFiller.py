import openpyxl
from docx import Document
import sys


def fill_table_from_excel(excel_filename, doc_filename, start_row, row_count):
    # Open the Excel workbook and select the active worksheet
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active

    # Load the Word document
    doc = Document(doc_filename)

    # Get the table in the document and the header row
    table = doc.tables[0]
    header_row = table.rows[0]

    # Get the positions of the temporary words in the header row
    tmp_word_positions = [cell.text for cell in header_row.cells if cell.text]
    tmp_excel_positions = [cell.value for cell in ws[1] if cell.value]

    # Loop through the non-empty rows in the Excel worksheet
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + row_count - 1, values_only=True):
        if any(row):
            # Add a new row to the table
            new_row = table.add_row()
            variables = {val: row[i] for i, val in enumerate(tmp_excel_positions)}

            # Loop through the cells in the new row and replace the temporary words with the data from Excel
            for i, xpos in enumerate(tmp_word_positions):
                # beautify kuotas
                clearxpos = xpos.replace(u'\u2018', "'").replace(u'\u2019', "'").replace(u'\u201C', '"').replace(u'\u201D', '"')
                z = eval(clearxpos, variables)
                new_value = z
                new_row.cells[i].text = new_value.replace(xpos, "").strip()

    # Save the modified Word document
    doc.save("0" + doc_filename)


if __name__ == '__main__':
    if len(sys.argv) < 5:
        print("Usage: python script.py excel_file doc_file start_row count_rows")
        sys.exit(1)
    fill_table_from_excel(sys.argv[1], sys.argv[2], int(sys.argv[3]), int(sys.argv[4]))
